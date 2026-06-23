"""
Excel report generation.

Builds the platform's required report format exactly:

    | File Name | Test Case No | Result |

using pandas + openpyxl, with light styling (bold header, conditional
PASS/FAIL cell coloring, autosized columns) since this is meant to be
opened directly by a QA engineer.
"""

from __future__ import annotations

import logging
from pathlib import Path

import pandas as pd
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from backend.executors.base_executor import ExecutionResult

logger = logging.getLogger(__name__)

_PASS_FILL = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")
_FAIL_FILL = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
_HEADER_FILL = PatternFill(start_color="305496", end_color="305496", fill_type="solid")
_HEADER_FONT = Font(color="FFFFFF", bold=True)


def generate_excel_report(results: list[ExecutionResult], output_path: Path) -> Path:
    """
    Write `results` to an .xlsx file at `output_path` in the required
    | File Name | Test Case No | Result | format, with TC numbers
    assigned sequentially in upload order.

    Returns the output path for convenience.
    """
    rows = []
    for index, result in enumerate(results, start=1):
        tc_number = f"TC{index:03d}"
        rows.append(
            {
                "File Name": result.filename,
                "Test Case No": tc_number,
                "Result": "PASS" if result.passed else "FAIL",
            }
        )

    df = pd.DataFrame(rows, columns=["File Name", "Test Case No", "Result"])

    output_path.parent.mkdir(parents=True, exist_ok=True)
    with pd.ExcelWriter(output_path, engine="openpyxl") as writer:
        df.to_excel(writer, index=False, sheet_name="Test Results")
        _apply_styling(writer, df)

    logger.info("Generated Excel report at '%s' with %d row(s)", output_path, len(rows))
    return output_path


def _apply_styling(writer: pd.ExcelWriter, df: pd.DataFrame) -> None:
    worksheet = writer.sheets["Test Results"]

    # Header styling
    for col_idx in range(1, len(df.columns) + 1):
        cell = worksheet.cell(row=1, column=col_idx)
        cell.fill = _HEADER_FILL
        cell.font = _HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Conditional PASS/FAIL coloring + column autosize
    result_col_idx = df.columns.get_loc("Result") + 1
    for row_idx in range(2, len(df) + 2):
        cell = worksheet.cell(row=row_idx, column=result_col_idx)
        cell.alignment = Alignment(horizontal="center")
        if cell.value == "PASS":
            cell.fill = _PASS_FILL
        elif cell.value == "FAIL":
            cell.fill = _FAIL_FILL

    for col_idx, column_name in enumerate(df.columns, start=1):
        max_len = max(
            [len(str(column_name))] + [len(str(v)) for v in df[column_name].tolist()] or [10]
        )
        worksheet.column_dimensions[get_column_letter(col_idx)].width = max_len + 6
