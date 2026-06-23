"""Unit tests for backend.core.excel_generator."""

from __future__ import annotations

from pathlib import Path

import openpyxl

from backend.core.excel_generator import generate_excel_report
from backend.executors.base_executor import ExecutionResult


def test_report_has_correct_columns_and_sequential_tc_numbers(tmp_path: Path) -> None:
    results = [
        ExecutionResult(
            filename="login_test.py",
            passed=True,
            stage="test",
            stdout="",
            stderr="",
            exit_code=0,
        ),
        ExecutionResult(
            filename="payment.py",
            passed=False,
            stage="test",
            stdout="",
            stderr="AssertionError",
            exit_code=1,
        ),
        ExecutionResult(
            filename="calc.c",
            passed=True,
            stage="run",
            stdout="",
            stderr="",
            exit_code=0,
        ),
    ]

    output_path = tmp_path / "report.xlsx"
    generate_excel_report(results, output_path)

    wb = openpyxl.load_workbook(output_path)
    ws = wb["Test Results"]

    header = [cell.value for cell in next(ws.iter_rows(min_row=1, max_row=1))]
    assert header == ["File Name", "Test Case No", "Result"]

    rows = list(ws.iter_rows(min_row=2, values_only=True))
    assert rows[0] == ("login_test.py", "TC001", "PASS")
    assert rows[1] == ("payment.py", "TC002", "FAIL")
    assert rows[2] == ("calc.c", "TC003", "PASS")
